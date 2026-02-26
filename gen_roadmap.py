
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
OUTPUT_PATH = r"C:\Users\Hp\OneDrive\Desktop\Roadmap\Roadmap.xlsx"

COLORS = {
    "header_bg":  "1A1A2E",
    "header_fg":  "E8EEFF",
    "border":     "CBD5E1",
}

ROW_COLORS = {
    "backend":   "E8F5E9",
    "uni":       "FFF3E0",
    "office":    "F3EAF8",
    "travel":    "FFFFF0",
    "recovery":  "E3F2FD",
    "sleep":     "E8EAF6",
    "reflect":   "FFF8E1",
    "planning":  "E0F7FA",
    "wind_down": "FCE4EC",
    "milestone": "FFD700",
    "break":     "F9F9F9",
    "default":   "FFFFFF",
}

PHASE_HEADER_COLORS = {
    1: "0D47A1",
    2: "4A148C",
    3: "1B5E20",
    4: "BF360C",
}

DAY_BANNER_COLORS = {
    1: "BBDEFB",  # Phase 1 – light blue
    2: "E1BEE7",  # Phase 2 – light purple
    3: "C8E6C9",  # Phase 3 – light green
    4: "FFE0B2",  # Phase 4 – light orange
}

WEEK_FOCUS = {
    1:  "Node.js Internals + HTTP Server",
    2:  "Express.js Deep + Routing + Middleware",
    3:  "PostgreSQL + Schema Design + Auth",
    4:  "JWT + Env Management + Security",
    5:  "Clean Architecture + MVC Pattern",
    6:  "Error Handling + Logging + Production Design",
    7:  "Docker + Containerization",
    8:  "CI/CD + GitHub Actions",
    9:  "Redis + Queues + Async Workers",
    10: "Rate Limiting + Caching + Performance",
    11: "Monitoring + Health Checks + PM2",
    12: "Capstone – Full Backend Blueprint",
    13: "Review + Polish + GitHub Portfolio + Reflection",
}

WEEK_MILESTONE = {
    1:  "Simple HTTP Server working without Express",
    2:  "CRUD API with clean routing structure",
    3:  "Auth API with JWT + PostgreSQL working",
    4:  "Secure Token Refresh System complete",
    5:  "MVC backend skeleton deployed to GitHub",
    6:  "Production-style error + logging system",
    7:  "Node API containerized with Docker",
    8:  "GitHub Actions CI/CD pipeline running",
    9:  "Background email queue worker functional",
    10: "Rate-limited + cached API live",
    11: "Health check + monitored API deployed",
    12: "Full backend system design document done",
    13: "90-Day GitHub portfolio + identity shift",
}

WEEK_ROLE = {
    1:  "🟢 Code Cadet",
    2:  "🟡 Route Commander",
    3:  "🟠 Data Architect",
    4:  "🔵 Security Operator",
    5:  "🔵 System Engineer",
    6:  "🔴 Stability Builder",
    7:  "🟤 Deployment Operator",
    8:  "⚫ Pipeline Architect",
    9:  "🟡 Async Specialist",
    10: "🔵 Performance Engineer",
    11: "🟣 Reliability Designer",
    12: "🔴 Production Architect",
    13: "🏆 Backend Engineer",
}

BACKEND_TASKS = {
    1:  [
        "Build simple HTTP server with Node.js (no Express) — handle GET /hello",
        "Explore Node.js event loop — write async code, understand callbacks",
        "Setup npm project, explore package.json, write first modular script",
        "Practice fs, path, os modules — read/write files programmatically",
        "Build a CLI mini-tool using Node.js process and readline",
        "Read Node.js official docs: Events, Streams, Timer APIs",
        "WEEK 1 MINI PROJECT: HTTP server with /hello /about /api routes + JSON response",
    ],
    2:  [
        "Install Express.js, setup basic app, understand req/res lifecycle",
        "Build RESTful CRUD routes: GET, POST, PUT, DELETE for /tasks",
        "Implement middleware chain: logging, body parser, custom middleware",
        "Practice route parameters (:id), query strings, nested routes",
        "Implement Express Router — split routes into separate files",
        "Add basic error middleware (4-param function) to Express app",
        "WEEK 2 MINI PROJECT: Todo CRUD API with Express + Router + Middleware",
    ],
    3:  [
        "Install PostgreSQL + pg npm module, connect to DB from Node.js",
        "Design schema: users + tasks tables with relationships",
        "Write raw SQL queries: SELECT, INSERT, UPDATE, DELETE with pg",
        "Intro to Prisma ORM — schema.prisma, migrate, basic CRUD",
        "Practice SQL JOINs — fetch user with their tasks in one query",
        "Link PostgreSQL DB to Express API — replace in-memory data",
        "WEEK 3 MINI PROJECT: Auth system — register, login, JWT + DB",
    ],
    4:  [
        "Deep dive JWT: jsonwebtoken — sign, verify, decode, expiry",
        "Setup dotenv — all secrets in .env, never hardcode values",
        "Build authMiddleware — protect routes with JWT verification",
        "Implement token refresh endpoint — refresh token pattern",
        "Add Helmet.js + CORS to Express — security headers setup",
        "Practice: protect a full API with auth middleware + role check",
        "WEEK 4 MINI PROJECT: Secure Auth system with refresh token + protected routes",
    ],
    5:  [
        "Study MVC pattern — Model, View (API response), Controller",
        "Refactor existing API: separate controllers from routes",
        "Add service layer — business logic NOT in controller",
        "Implement repository pattern — DB queries in repository files",
        "Add config management — all env vars accessed from config.js",
        "Practice: refactor Todo API to full MVC + Service + Repo structure",
        "WEEK 5 MINI PROJECT: Clean MVC backend skeleton pushed to GitHub",
    ],
    6:  [
        "Setup Winston logger — log levels: info, warn, error",
        "Build global error handler middleware — catch all errors centrally",
        "Define custom error classes: ValidationError, AuthError, NotFoundError",
        "Practice HTTP status codes map — 200, 201, 400, 401, 403, 404, 500",
        "Add request logging middleware — log every incoming request",
        "Write logs to rotating file system (winston-daily-rotate-file)",
        "WEEK 6 MINI PROJECT: Production-style API with centralized logger + error handler",
    ],
    7:  [
        "Install Docker Desktop, understand images vs containers",
        "Write Dockerfile for Node.js app — FROM, WORKDIR, COPY, CMD",
        "Build and run Docker image locally — docker build + docker run",
        "Setup docker-compose.yml — Node app + PostgreSQL together",
        "Understand Docker volumes — persist DB data across restarts",
        "Push image to Docker Hub — docker push basics",
        "WEEK 7 MINI PROJECT: Full API containerized with docker-compose (Node + DB)",
    ],
    8:  [
        "Understand GitHub Actions: workflows, jobs, steps, triggers",
        "Create CI workflow: on push to main — install + test",
        "Add Jest unit tests — write 3 tests for your API endpoints",
        "Integrate tests in GitHub Actions CI pipeline",
        "Understand CD concept — deploy to server on merge to main",
        "Study GitHub secrets — store ENV variables securely in CI",
        "WEEK 8 MINI PROJECT: CI/CD pipeline running on GitHub for your API",
    ],
    9:  [
        "Install Redis — understand key-value store concept",
        "Install BullMQ — create a simple queue and add jobs",
        "Build a worker process — consume jobs from queue",
        "Implement email queue: add email job when user registers",
        "Add retry logic + failure handling to BullMQ jobs",
        "Monitor queues with Bull Board (admin UI for BullMQ)",
        "WEEK 9 MINI PROJECT: Email notification queue worker with BullMQ + Redis",
    ],
    10: [
        "Setup express-rate-limit — rate limit by IP",
        "Implement Redis caching — cache API responses with TTL",
        "Understand cache invalidation — when to clear cache",
        "Test API response time with/without cache using Postman",
        "Add compression middleware — reduce response size",
        "Apply caching to heavy DB query endpoint",
        "WEEK 10 MINI PROJECT: Rate-limited + Redis-cached API with performance comparison",
    ],
    11: [
        "Add /health endpoint — return server status + DB connection status",
        "Setup PM2 — process manager for Node.js in production",
        "Configure PM2 ecosystem.config.js — cluster mode",
        "Implement graceful shutdown — handle SIGTERM cleanly",
        "Study log aggregation concept — centralized logging approach",
        "Add uptime monitoring concept (Better Uptime or simple ping check)",
        "WEEK 11 MINI PROJECT: Monitored + stable API with PM2 + health check endpoint",
    ],
    12: [
        "Study load balancing — horizontal scaling concept",
        "Document your full backend system design in a README",
        "Explore database indexing — add indexes to slow queries",
        "Study API versioning — /api/v1 vs /api/v2 pattern",
        "Compare monolith vs microservices — when to use what",
        "Design a scalable e-commerce backend architecture (diagram)",
        "CAPSTONE: Full backend system design document + architecture diagram on GitHub",
    ],
    13: [
        "Review all 12 weeks — fill any gaps in understanding",
        "Go through each project repo — add proper README files",
        "Fix incomplete or buggy projects from earlier weeks",
        "Polish GitHub profile — pin 3 best backend projects",
        "Write a personal '90-Day Backend Journey' blog on Dev.to or GitHub",
        "Plan next 90 days — what comes after this phase",
        "FINAL: 90-Day identity shift documented. Portfolio ready. Next roadmap set.",
    ],
}

AUDIO_TOPICS = {
    1: "Node.js fundamentals: event loop, async, modules",
    2: "Express.js: routing, middleware, MVC basics",
    3: "Databases: SQL, PostgreSQL, ORMs explained",
    4: "JWT, OAuth, security in web APIs",
    5: "Software architecture: MVC, Clean Code, SOLID",
    6: "Production backend: logging, error handling, monitoring",
    7: "Docker and containerization for developers",
    8: "CI/CD, DevOps basics, GitHub Actions",
    9: "Redis, queues, async background jobs",
    10: "Caching strategies and performance optimization",
    11: "SRE, monitoring, reliability engineering basics",
    12: "System design for backend engineers",
    13: "Backend career growth and learning strategy",
}

UNI_MOTIVATION = {
    "TOA": "Automata is the mathematical foundation of all computing. Understand it deeply.",
    "SE":  "Software Engineering teaches you to build systems — not just features.",
    "PD":  "Personal Development is the software upgrade for your mindset.",
    "IS":  "Information Security is the shield that every backend engineer must master.",
    "OS":  "Operating Systems: understand what runs beneath your code.",
    "FA":  "Financial Accounting: every founder must understand numbers.",
}

MOTIVATIONS = {
    "planning":  "Every master starts with a clear plan.",
    "travel":    "Even the road is a classroom — ears open.",
    "office":    "Deep output > long hours. Focus wins.",
    "backend":   "Hands-on beats theory. Build it, break it, fix it.",
    "uni":       "Absorb, summarize, retain. One class at a time.",
    "study":     "Every solved problem = confidence built.",
    "dinner":    "Balance keeps the engine running long-term.",
    "project":   "Small bricks today, production system tomorrow.",
    "reflect":   "Reflection turns effort into mastery.",
    "wind":      "Feed curiosity before sleep.",
    "sleep":     "Sleep is the reset button for brilliance.",
    "break":     "Short reset = longer, sharper focus.",
    "recovery":  "Recovery is not laziness. It is strategy.",
    "sunday":    "Full recharge today = unstoppable week ahead.",
    "milestone": "Milestone reached. Proof that the system works.",
    "cortellect":"Lead, don't do. Delegate to scale.",
    "badar":     "Revenue enables the learning. Protect it.",
}

def get_phase(week):
    if week <= 3: return 1
    if week <= 6: return 2
    if week <= 9: return 3
    return 4

def day_name(day_index):
    return ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"][day_index]

def build_day(week, abs_day, day_idx):
    """Returns list of row dicts for one day."""
    dname = day_name(day_idx)
    wf    = WEEK_FOCUS[week]
    wr    = WEEK_ROLE[week]
    bt    = BACKEND_TASKS[week][day_idx]
    at    = AUDIO_TOPICS[week]
    phase = get_phase(week)

    rows = []
    def r(time, topic, task, role, motive, ref, cat):
        rows.append(dict(
            week=week, day=abs_day, day_name=dname,
            time=time, topic=topic, task=task, role=role,
            progress=0, status="Not Started",
            motivation=motive, ref=ref, notes="–", cat=cat
        ))

    # ──────────── SUNDAY ────────────
    if dname == "Sunday":
        r("07:30–08:00 AM", "Wake Up + Light Breakfast", "Hydrate, light stretch, disconnect from work", "🌅 Rest Architect", MOTIVATIONS["recovery"], "–", "recovery")
        r("08:00–09:00 AM", "Morning Walk", "30–40 min walk. No phone. Just breathe.", "🌿 Energy Guardian", MOTIVATIONS["sunday"], "–", "recovery")
        r("09:00–10:30 AM", "Weekly Review", f"Review Week {week}: what worked, what to improve", "🧭 Weekly Strategist", MOTIVATIONS["reflect"], "Notion / Notebook", "reflect")
        r("10:30–12:30 PM", "Light University Revision", "Review 1 subject lightly. No deep study today.", "📖 Academic Keeper", MOTIVATIONS["uni"], "LMS / Notes", "uni")
        r("12:30–01:30 PM", "Lunch + Family Time", "Eat + family time. Full offline.", "🌿 Balanced Achiever", MOTIVATIONS["dinner"], "–", "recovery")
        r("01:30–03:00 PM", f"Backend Concept Recap — {wf}", f"Re-read key notes from this week's backend study. Let it sink.", "💡 Knowledge Consolidator", "Review what you built. Let the learning consolidate.", "Notes / GitHub", "backend")
        r("03:00–04:30 PM", "Next Week Planning", f"Plan Week {min(week+1,13)} backend + university tasks", "🗓 Weekly Planner", "Precise planning = stress-free week ahead.", "Roadmap / Notion", "reflect")
        r("04:30–06:00 PM", "Hobbies / Complete Rest", "Free time only. No tech pressure.", "🎮 Human Being", MOTIVATIONS["recovery"], "–", "recovery")
        r("06:00–07:30 PM", "Dinner + Relax", "Family + food. Full disconnect.", "🌿 Balanced Achiever", MOTIVATIONS["dinner"], "–", "recovery")
        r("07:30–09:00 PM", "Light Tech Reading (Optional)", "1 article or short YouTube video max. No pressure.", "📚 Mind Explorer", MOTIVATIONS["wind"], "Dev.to / YouTube", "wind_down")
        r("09:00–10:00 PM", "Prepare for Monday", "Set VSCode ready, check assignments, plan Monday", "🗓 Monday Architect", "Prepare the battlefield the night before.", "Notion / Roadmap", "reflect")
        r("10:00–10:30 PM", "Wind Down", "Calm music, no screens 30 min before sleep", "🌙 Night Scholar", MOTIVATIONS["wind"], "Book / Music", "wind_down")
        r("10:30 PM–07:00 AM", "Sleep — 8 Hours", "Full cognitive reset. Non-negotiable.", "🛌 Energy Guardian", MOTIVATIONS["sleep"], "–", "sleep")
        # Milestone on Day 7 of each week
        if day_idx == 6:
            rows.append(dict(
                week=week, day=abs_day, day_name=dname,
                time="🏆 WEEK MILESTONE",
                topic=f"Week {week} Milestone: {WEEK_MILESTONE[week]}",
                task=f"Complete and push: {WEEK_MILESTONE[week]}. Document. Commit to GitHub.",
                role=wr, progress=0, status="Not Started",
                motivation=MOTIVATIONS["milestone"], ref="GitHub / Notion",
                notes="Milestone complete = proof the system works.", cat="milestone"
            ))
        return rows

    # ──────────── WEEKDAY COMMON START ────────────
    r("07:00–07:30 AM", "Wake Up + Morning Routine", "Freshen up, stretch, hydrate. No phone for 15 min.", "⏰ Morning Starter", "Your morning routine is your performance foundation.", "–", "recovery")
    r("07:30–08:00 AM", "Breakfast + Daily Planning", "Top 3 tasks of the day. Check uni + backend goals.", "🌱 Initiator", MOTIVATIONS["planning"], "Notion / Notebook", "planning")
    r("08:00–08:45 AM", "Travel (Home → Office / University)", f"Audio learning: {at}", "🎧 Knowledge Commuter", MOTIVATIONS["travel"], "Airbuds + Podcast/Audio", "travel")

    # ──────────── MONDAY ────────────
    if dname == "Monday":
        r("09:00–12:30 PM", "Office Work — Deep Focus", "Deep focus: 1 major task + general office work", "💼 Office Strategist", MOTIVATIONS["office"], "VSCode / GitHub", "office")
        r("12:30–01:00 PM", "Lunch / Recharge", "Eat + short walk before uni commute", "🌿 Energy Keeper", "Refuel — TOA class needs concentrated energy.", "–", "recovery")
        r("01:00–01:30 PM", "Travel Office → University", f"Audio: {at}", "🎧 Road Scholar", MOTIVATIONS["travel"], "Airbuds", "travel")
        r("01:30–03:00 PM", "University — Theory of Automata (TOA)", "Active note-taking, follow examples, ask questions", "🎓 Student Coder", UNI_MOTIVATION["TOA"], "LMS / Class Notes", "uni")
        r("03:00–03:45 PM", "Travel University → Home", "Audio recap — reinforce TOA concepts", "🎧 Road Scholar", "Concepts consolidate during commute.", "Audio Notes", "travel")
        r("03:45–04:00 PM", "Micro Break", "Stretch, hydrate, decompress", "🧘 Micro Reset", MOTIVATIONS["break"], "–", "recovery")
        r("04:00–06:00 PM", f"Backend — {wf}", bt, wr, MOTIVATIONS["backend"], "VSCode / GitHub / Docs", "backend")
        r("06:00–06:30 PM", "Tea Break", "Refresh, snack", "🌿 Energy Keeper", "Recharge for evening sprint.", "–", "break")
        r("06:30–08:00 PM", "University Study — TOA", "Review notes, solve exercises, prep quiz/assignment", "📝 Academic Builder", MOTIVATIONS["study"], "LMS / Notebook", "uni")
        r("08:00–09:00 PM", "Dinner + Family Time", "Eat, full disconnect from work", "🌿 Balanced Achiever", MOTIVATIONS["dinner"], "–", "recovery")
        r("09:00–10:30 PM", f"Backend Mini Project — Week {week}", bt, "🧱 System Builder", MOTIVATIONS["project"], "VSCode / GitHub", "backend")
        r("10:30–11:00 PM", "Reflection + Plan Tuesday", "3 wins today. 1 improvement. Plan tomorrow.", "🌟 Daily Strategist", MOTIVATIONS["reflect"], "Notion / Notebook", "reflect")
        r("11:00–11:30 PM", "Wind Down", "Light reading. No screens 30 min before sleep.", "📚 Mind Explorer", MOTIVATIONS["wind"], "Book / Dev.to", "wind_down")
        r("11:30 PM–07:00 AM", "Sleep — 7.5 Hours", "Non-negotiable. Brain consolidates code while you sleep.", "🛌 Energy Guardian", MOTIVATIONS["sleep"], "–", "sleep")

    # ──────────── TUESDAY ────────────
    elif dname == "Tuesday":
        r("08:45–09:15 AM", "Travel Home → University (early for SE)", f"Audio: {at}", "🎧 Knowledge Commuter", MOTIVATIONS["travel"], "Airbuds", "travel")
        r("09:15–10:30 AM", "Office Work — Mini Session", "1 focused backend task or quick office work", "💼 Office Strategist", MOTIVATIONS["office"], "VSCode / GitHub", "office")
        r("10:30–12:00 PM", "University — Software Engineering (SE)", "Active note-taking, participate, summarize each segment", "🎓 Student Coder", UNI_MOTIVATION["SE"], "LMS / Class Notes", "uni")
        r("12:00–12:45 PM", "Lunch / Recharge", "Eat + short walk", "🌿 Energy Keeper", "Fuel body, power mind.", "–", "recovery")
        r("12:45–02:30 PM", "Office Work — Deep Focus", "Major deep work session: core office task", "💼 Office Strategist", MOTIVATIONS["office"], "VSCode / GitHub", "office")
        r("02:30–04:00 PM", f"Backend — {wf}", bt, wr, MOTIVATIONS["backend"], "VSCode / GitHub / Docs", "backend")
        r("04:00–04:45 PM", "Travel Home", "Audio recap — reinforce backend concepts", "🎧 Road Scholar", MOTIVATIONS["travel"], "Audio Notes", "travel")
        r("04:45–05:00 PM", "Micro Break", "Stretch / hydrate", "🧘 Micro Reset", MOTIVATIONS["break"], "–", "recovery")
        r("05:00–06:30 PM", "Cortellect Supervision (No Coding)", "Review team tasks, approve outputs, delegate only", "🏢 Team Lead", MOTIVATIONS["cortellect"], "Slack / Notion", "office")
        r("06:30–07:30 PM", "University Study — SE", "Review SE notes, solve exercises", "📝 Academic Builder", MOTIVATIONS["study"], "LMS / Notes", "uni")
        r("07:30–08:30 PM", "Dinner + Family Time", "Full offline, family present", "🌿 Balanced Achiever", MOTIVATIONS["dinner"], "–", "recovery")
        r("08:30–10:00 PM", f"Backend Mini Project — Week {week}", bt, "🧱 System Builder", MOTIVATIONS["project"], "VSCode / GitHub", "backend")
        r("10:00–10:30 PM", "Reflection + Plan Wednesday", "3 wins + 1 improvement. Tomorrow is heavy — plan it.", "🌟 Daily Strategist", MOTIVATIONS["reflect"], "Notion / Notebook", "reflect")
        r("10:30–11:00 PM", "Wind Down", "Light reading, calm down", "📚 Mind Explorer", MOTIVATIONS["wind"], "Book", "wind_down")
        r("11:00 PM–07:00 AM", "Sleep EARLY — Heavy Day Tomorrow", "Sleep early before heavy Wednesday (PD + IS = 4.5 hrs class)", "🛌 Energy Guardian", "Sleep early = perform on heavy days.", "–", "sleep")

    # ──────────── WEDNESDAY ────────────
    elif dname == "Wednesday":
        r("08:00–10:00 AM", f"Morning Deep Work — {wf}", bt, wr, MOTIVATIONS["backend"], "VSCode / GitHub", "backend")
        r("10:00–10:30 AM", "Break + Prep for Uni", "Light snack, review PD lecture topic", "🌿 Energy Keeper", MOTIVATIONS["break"], "LMS Notes", "recovery")
        r("10:30–11:15 AM", "Travel Home → University", f"Audio: {at}", "🎧 Road Scholar", MOTIVATIONS["travel"], "Airbuds", "travel")
        r("11:15–12:00 PM", "Arrive Uni + Canteen Lunch", "Eat before 4.5 hour class block", "🌿 Energy Keeper", "Fuel before the marathon class block.", "–", "recovery")
        r("12:00–03:00 PM", "University — Personal Development (PD)", "Active notes, reflect on each concept, participate", "🎓 Knowledge Seeker", UNI_MOTIVATION["PD"], "LMS / Class Notes", "uni")
        r("03:00–04:30 PM", "University — Information Security (IS)", "Active notes, practice exercises, mini quiz prep", "🔒 Security Explorer", UNI_MOTIVATION["IS"], "LMS / Class Notes", "uni")
        r("04:30–05:15 PM", "Travel Uni → Home", "Audio recap: IS + PD concepts", "🎧 Road Scholar", "Even the ride home reinforces learning.", "Audio Notes", "travel")
        r("05:15–05:30 PM", "Micro Break", "Stretch, hydrate, decompress", "🧘 Micro Reset", "4.5 hours of class done. You earned this break.", "–", "recovery")
        r("05:30–06:30 PM", "Light University Study", "Review PD or IS notes gently — no deep work", "📝 Academic Builder", "Consolidate today's learning gently.", "LMS / Notes", "uni")
        r("06:30–07:30 PM", "Dinner + Family Time", "Full offline. Brain recovery mode.", "🌿 Balanced Achiever", MOTIVATIONS["dinner"], "–", "recovery")
        r("07:30–08:00 PM", "Plan Thursday + Weekly Check", "Quick 30-min review: backend progress + assignments", "🧭 Weekly Strategist", "Stay one step ahead by planning ahead.", "Notion / Notebook", "reflect")
        r("08:00–08:30 PM", "Wind Down", "Light reading, calm music, dark room", "📚 Mind Explorer", MOTIVATIONS["wind"], "Book / Music", "wind_down")
        r("08:30 PM–07:00 AM", "FULL RECOVERY SLEEP — 10.5 hrs", "Wednesday is the heaviest day. Sleep fully. No exception.", "🛌 Energy Guardian", "Heavy days demand complete recovery nights.", "–", "sleep")

    # ──────────── THURSDAY ────────────
    elif dname == "Thursday":
        r("08:00–10:00 AM", f"Morning Deep Work — {wf}", bt, wr, MOTIVATIONS["backend"], "VSCode / GitHub", "backend")
        r("10:00–10:30 AM", "Break + TOA Prep", "Snack, review TOA notes from Monday", "🌿 Energy Keeper", MOTIVATIONS["break"], "LMS Notes", "recovery")
        r("10:30–12:30 PM", "Office Work — Deep Focus", "Complete 1 major office task or backend module", "💼 Office Strategist", MOTIVATIONS["office"], "VSCode / GitHub", "office")
        r("12:30–01:00 PM", "Lunch / Recharge", "Eat healthy — long afternoon ahead", "🌿 Energy Keeper", "Fuel before 4.5 hours in university.", "–", "recovery")
        r("01:00–01:30 PM", "Travel Office → University", f"Audio: {at}", "🎧 Road Scholar", MOTIVATIONS["travel"], "Airbuds", "travel")
        r("01:30–03:00 PM", "University — Theory of Automata (TOA)", "Active notes, solve exercises, ask questions", "🎓 Student Coder", UNI_MOTIVATION["TOA"], "LMS / Class Notes", "uni")
        r("03:00–06:00 PM", "University — Operating Systems Lab", "Lab exercises, hands-on practicals, document steps", "🖥 Lab Engineer", UNI_MOTIVATION["OS"], "Lab PC / Lab Manual", "uni")
        r("06:00–06:45 PM", "Travel Uni → Home", "Decompress — light audio or silence on the way", "🎧 Road Scholar", "Long day done. Just breathe on the ride home.", "Airbuds (optional)", "travel")
        r("06:45–07:30 PM", "Dinner + Full Rest", "Eat + family time. No screens.", "🌿 Balanced Achiever", MOTIVATIONS["dinner"], "–", "recovery")
        r("07:30–08:30 PM", "University Study — Light", "Review OS Lab notes only — max 1 hour, keep it light", "📝 Academic Builder", "Consolidate lab concepts while still fresh.", "LMS / Notes", "uni")
        r("08:30–09:00 PM", "Plan Friday", "Plan: SE class + backend + Badar revenue task", "🧭 Daily Strategist", "Prepare Friday the night before.", "Notion / Notebook", "reflect")
        r("09:00–09:30 PM", "Wind Down", "Calm music, no work talk", "📚 Mind Explorer", MOTIVATIONS["wind"], "Book / Music", "wind_down")
        r("09:30 PM–07:00 AM", "FULL RECOVERY SLEEP — 9.5 hrs", "Thursday is heavy. Complete sleep mandatory.", "🛌 Energy Guardian", "Heavy days demand heavy recovery.", "–", "sleep")

    # ──────────── FRIDAY ────────────
    elif dname == "Friday":
        r("08:30–09:15 AM", "Travel Home → University (SE class)", f"Audio: {at}", "🎧 Knowledge Commuter", MOTIVATIONS["travel"], "Airbuds", "travel")
        r("09:15–10:30 AM", "Office Work — Quick Deep Focus", "1 important office task completed before SE class", "💼 Office Strategist", MOTIVATIONS["office"], "VSCode / GitHub", "office")
        r("10:30–12:00 PM", "University — Software Engineering (SE)", "Active notes, project discussion, apply from backend", "🎓 Student Coder", UNI_MOTIVATION["SE"], "LMS / Class Notes", "uni")
        r("12:00–12:45 PM", "Lunch / Recharge", "Eat + short walk", "🌿 Energy Keeper", "Fuel up for afternoon backend deep session.", "–", "recovery")
        r("12:45–02:30 PM", f"Backend Deep Session — {wf}", bt, wr, MOTIVATIONS["backend"], "VSCode / GitHub / Docs", "backend")
        r("02:30–04:00 PM", "Office Work — Badar Revenue Task", "Most critical Badar revenue task of the week", "💼 Office Strategist", MOTIVATIONS["badar"], "VSCode / Client tools", "office")
        r("04:00–04:45 PM", "Travel Home", "Audio recap", "🎧 Road Scholar", MOTIVATIONS["travel"], "Audio Notes", "travel")
        r("04:45–05:00 PM", "Micro Break", "Stretch / hydrate / short walk", "🧘 Micro Reset", MOTIVATIONS["break"], "–", "recovery")
        r("05:00–06:00 PM", f"Backend Mini Project — Week {week}", bt, "🧱 System Builder", MOTIVATIONS["project"], "VSCode / GitHub", "backend")
        r("06:00–07:00 PM", "University Study — SE", "SE notes + exercises + upcoming quiz prep", "📝 Academic Builder", MOTIVATIONS["study"], "LMS / Notes", "uni")
        r("07:00–08:00 PM", "Dinner + Family Time", "Full offline, family time", "🌿 Balanced Achiever", MOTIVATIONS["dinner"], "–", "recovery")
        r("08:00–09:00 PM", "Cortellect / Badar Weekly Check", "Team review, approve outputs, plan next week delegation", "🏢 CEO Mode", MOTIVATIONS["cortellect"], "Slack / Notion", "office")
        r("09:00–09:30 PM", "Reflection + Weekend Plan", f"Review Week {week} progress. Plan Saturday + Sunday.", "🌟 Weekly Planner", MOTIVATIONS["reflect"], "Notion / Notebook", "reflect")
        r("09:30–10:00 PM", "Wind Down", "Light reading / calm content", "📚 Mind Explorer", MOTIVATIONS["wind"], "Book / Dev.to", "wind_down")
        r("10:00 PM–07:00 AM", "Sleep — Full 9 Hours", "Recover fully — Saturday is also heavy.", "🛌 Energy Guardian", MOTIVATIONS["sleep"], "–", "sleep")

    # ──────────── SATURDAY ────────────
    elif dname == "Saturday":
        r("07:00–08:00 AM", "Wake Up + Planning", "Breakfast + review weekend goals. Light start.", "🌱 Initiator", "Weekend focused work = week ahead advantage.", "Notion / Notebook", "planning")
        r("08:00–10:00 AM", f"Backend — {wf} (Peak Morning)", bt, wr, MOTIVATIONS["backend"], "VSCode / GitHub / Docs", "backend")
        r("10:00–10:30 AM", "Break + Uni Prep", "Snack, glance at FA notes briefly", "🌿 Energy Keeper", MOTIVATIONS["break"], "FA Notes", "recovery")
        r("10:30–11:15 AM", "Travel Home → University (FA+IS)", f"Audio: {at}", "🎧 Road Scholar", MOTIVATIONS["travel"], "Airbuds", "travel")
        r("11:15–12:00 PM", "Arrive Uni + Canteen Lunch", "Eat before 4.5 hour class block", "🌿 Energy Keeper", "Fuel before the class marathon.", "–", "recovery")
        r("12:00–03:00 PM", "University — Financial Accounting (FA)", "Active notes, solve exercises, mini quiz prep", "🎓 Accounting Scholar", UNI_MOTIVATION["FA"], "LMS / Class Notes", "uni")
        r("03:00–04:30 PM", "University — Information Security (IS)", "Active notes, practice exercises", "🔒 Security Explorer", UNI_MOTIVATION["IS"], "LMS / Class Notes", "uni")
        r("04:30–05:15 PM", "Travel Uni → Home", "Audio recap: IS + FA concepts", "🎧 Road Scholar", "Reinforce during the ride home.", "Audio Notes", "travel")
        r("05:15–05:30 PM", "Micro Break", "Hydrate, decompress", "🧘 Micro Reset", MOTIVATIONS["break"], "–", "recovery")
        r("05:30–06:30 PM", "University Study (Light)", "Review FA or IS notes — consolidate gently", "📝 Academic Builder", "Absorb what was taught. Don't force it.", "LMS / Notes", "uni")
        r("06:30–07:30 PM", "Dinner + Family Time", "Offline completely.", "🌿 Balanced Achiever", MOTIVATIONS["dinner"], "–", "recovery")
        r("07:30–09:00 PM", f"Backend Mini Project — Week {week} (Weekend Push)", bt, "🧱 System Builder", MOTIVATIONS["project"], "VSCode / GitHub", "backend")
        r("09:00–09:30 PM", "Pending Assignments", "Quick catch-up on any pending assignment", "📝 Academic Builder", MOTIVATIONS["study"], "LMS / Notes", "uni")
        r("09:30–10:00 PM", "Reflection — Saturday", "Note wins + 1 improvement. Update progress tracker.", "🌟 Daily Strategist", MOTIVATIONS["reflect"], "Notion / Notebook", "reflect")
        r("10:00–10:30 PM", "Wind Down", "Light reading, calm music", "📚 Mind Explorer", MOTIVATIONS["wind"], "Book / Dev.to", "wind_down")
        r("10:30 PM–07:30 AM", "FULL RECOVERY SLEEP — BRAIN OFF", "Saturday is heavy. Evening OFF. Sleep fully.", "🛌 Energy Guardian", MOTIVATIONS["sleep"], "–", "sleep")

    return rows


def style_cell(cell, fill_color, bold=False, font_size=9, align="left", wrap=True, color="1A1A2E"):
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(bold=bold, size=font_size, name="Calibri", color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    thin = Side(style="thin", color=COLORS["border"])
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "90-Day Roadmap"

    # Column widths
    widths = [7, 6, 11, 17, 32, 40, 22, 10, 13, 42, 24, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # HEADER ROW
    headers = [
        "Week", "Day", "Day Name", "Time Slot", "Topic / Focus Area",
        "Task / Practical Challenge", "Role / Designation",
        "Progress %", "Status", "Motivation Line",
        "Reference / Resource", "Notes / Reflection"
    ]
    ws.row_dimensions[1].height = 38
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=COLORS["header_fg"], size=11, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="medium", color="4A90D9")
        cell.border = Border(bottom=thin, right=Side(style="thin", color=COLORS["border"]))

    row_num = 2
    current_phase = None

    for week in range(1, 14):
        phase = get_phase(week)

        # Phase divider
        if phase != current_phase:
            current_phase = phase
            phase_names = {
                1: "PHASE 1 – Foundation Control (Weeks 1–3): Node.js + Express + PostgreSQL + Auth",
                2: "PHASE 2 – System Thinking (Weeks 4–6): JWT + Clean Architecture + Logging",
                3: "PHASE 3 – Production Engineering (Weeks 7–9): Docker + CI/CD + Redis + Queues",
                4: "PHASE 4 – Controlled Architecture (Weeks 10–13): Cache + Monitor + Capstone + Portfolio",
            }
            ws.row_dimensions[row_num].height = 30
            ws.merge_cells(f"A{row_num}:L{row_num}")
            cell = ws.cell(row=row_num, column=1)
            cell.value = f"◆  {phase_names[phase]}  ◆"
            cell.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
            cell.fill = PatternFill("solid", fgColor=PHASE_HEADER_COLORS[phase])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

        for day_idx in range(7):
            abs_day = (week - 1) * 7 + day_idx + 1
            dname_str = day_name(day_idx)

            # Day banner
            ws.row_dimensions[row_num].height = 22
            ws.merge_cells(f"A{row_num}:L{row_num}")
            cell = ws.cell(row=row_num, column=1)
            banner_suffix = "  ★ REST & RECHARGE DAY ★" if dname_str == "Sunday" else ""
            cell.value = f"  📅  Week {week}  |  Day {abs_day}  |  {dname_str}{banner_suffix}  |  Focus: {WEEK_FOCUS[week]}"
            banner_color = "FCE4EC" if dname_str == "Sunday" else DAY_BANNER_COLORS[phase]
            cell.fill = PatternFill("solid", fgColor=banner_color)
            cell.font = Font(bold=True, size=9, name="Calibri", color="1A1A2E")
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            row_num += 1

            # Data rows
            schedule = build_day(week, abs_day, day_idx)
            for entry in schedule:
                ws.row_dimensions[row_num].height = 32
                rc = ROW_COLORS.get(entry["cat"], ROW_COLORS["default"])

                vals = [
                    entry["week"], str(entry["day"]), entry["day_name"],
                    entry["time"], entry["topic"], entry["task"],
                    entry["role"], f"{entry['progress']}%", entry["status"],
                    entry["motivation"], entry["ref"], entry["notes"]
                ]
                is_milestone = entry["cat"] == "milestone"
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_num, column=col, value=val)
                    style_cell(
                        cell, rc,
                        bold=is_milestone,
                        font_size=9 if not is_milestone else 10,
                        align="center" if col in [1, 2, 8, 9] else "left",
                        color="1A1A2E" if not is_milestone else "7B3F00"
                    )
                row_num += 1

            # Gap between days
            ws.row_dimensions[row_num].height = 8
            ws.merge_cells(f"A{row_num}:L{row_num}")
            ws.cell(row=row_num, column=1).fill = PatternFill("solid", fgColor="F8F9FA")
            row_num += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:L1"

    wb.save(OUTPUT_PATH)
    print(f"SUCCESS: Roadmap saved to {OUTPUT_PATH}")
    print(f"Total rows: {row_num}")


if __name__ == "__main__":
    create_excel()
